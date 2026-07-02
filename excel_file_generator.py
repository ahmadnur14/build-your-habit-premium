"""
Excel File Generator
Generates actual Excel files that can be downloaded
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelFileGenerator:
    """Generate actual Excel files (.xlsx format)"""
    
    def __init__(self, output_dir='downloads'):
        """Initialize Excel file generator"""
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_habit_tracker_template(self, filename='Habit_Tracker_Template.xlsx'):
        """Generate habit tracker template Excel file"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Dashboard sheet
        ws_dashboard = wb.create_sheet('Dashboard')
        self._create_dashboard_sheet(ws_dashboard)
        
        # Trackers sheet
        ws_trackers = wb.create_sheet('Trackers')
        self._create_trackers_sheet(ws_trackers)
        
        # Daily Log sheet
        ws_daily = wb.create_sheet('Daily Log')
        self._create_daily_log_sheet(ws_daily)
        
        # Statistics sheet
        ws_stats = wb.create_sheet('Statistics')
        self._create_statistics_sheet(ws_stats)
        
        # Goals sheet
        ws_goals = wb.create_sheet('Goals')
        self._create_goals_sheet(ws_goals)
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def generate_habit_data_export(self, tracker_data, filename='Habit_Data_Export.xlsx'):
        """Generate habit data export Excel file"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Overview sheet
        ws_overview = wb.create_sheet('Overview')
        self._create_overview_sheet(ws_overview, tracker_data)
        
        # All Habits sheet
        ws_habits = wb.create_sheet('All Habits')
        self._create_all_habits_data_sheet(ws_habits, tracker_data)
        
        # Statistics sheet
        ws_stats = wb.create_sheet('Statistics')
        self._create_statistics_data_sheet(ws_stats, tracker_data)
        
        # Insights sheet
        ws_insights = wb.create_sheet('Insights')
        self._create_insights_sheet(ws_insights)
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def generate_weekly_report(self, week_data=None, filename='Weekly_Report.xlsx'):
        """Generate weekly report Excel file"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet('Summary')
        self._create_weekly_summary(ws_summary)
        
        # Daily Details sheet
        ws_details = wb.create_sheet('Daily Details')
        self._create_weekly_details(ws_details)
        
        # Performance sheet
        ws_performance = wb.create_sheet('Performance')
        self._create_performance_sheet(ws_performance)
        
        # Goals Progress sheet
        ws_goals = wb.create_sheet('Goals Progress')
        self._create_goals_progress_sheet(ws_goals)
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def generate_monthly_report(self, month_data=None, filename='Monthly_Report.xlsx'):
        """Generate monthly report Excel file"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet('Summary')
        self._create_monthly_summary(ws_summary)
        
        # Daily Log sheet
        ws_daily = wb.create_sheet('Daily Log')
        self._create_monthly_daily_log(ws_daily)
        
        # Statistics sheet
        ws_stats = wb.create_sheet('Statistics')
        self._create_monthly_statistics(ws_stats)
        
        # Trends sheet
        ws_trends = wb.create_sheet('Trends')
        self._create_trends_sheet(ws_trends)
        
        # Achievements sheet
        ws_achievements = wb.create_sheet('Achievements')
        self._create_achievements_sheet(ws_achievements)
        
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
    
    def _create_dashboard_sheet(self, ws):
        """Create dashboard sheet with metrics"""
        ws['A1'] = 'Habit Tracker Dashboard'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['Metric', 'Value', 'Target', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        data = [
            ['Total Habits', 0, 0, 'N/A'],
            ['Active Habits', 0, 0, 'N/A'],
            ['Completion Rate (%)', 0, 100, 'Pending'],
            ['Current Streak', 0, 0, 'Active'],
            ['Average Score', 0, 80, 'Pending']
        ]
        
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_trackers_sheet(self, ws):
        """Create trackers sheet"""
        ws['A1'] = 'All Trackers'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        headers = ['Habit Name', 'Frequency', 'Unit', 'Status', 'Created Date', 'Total Entries']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_daily_log_sheet(self, ws):
        """Create daily log sheet"""
        ws['A1'] = 'Daily Log'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        headers = ['Date', 'Habit', 'Value', 'Unit', 'Notes', 'Time']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_statistics_sheet(self, ws):
        """Create statistics sheet"""
        ws['A1'] = 'Statistics'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        headers = ['Habit', 'Total Entries', 'Average', 'Max', 'Min', 'Completion %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_goals_sheet(self, ws):
        """Create goals sheet"""
        ws['A1'] = 'Goals'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        headers = ['Goal', 'Target', 'Progress', 'Deadline', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_overview_sheet(self, ws, tracker_data):
        """Create overview sheet with data"""
        ws['A1'] = 'Habit Tracking Overview'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        ws['A3'] = 'Total Habits:'
        ws['B3'] = len(tracker_data)
        
        ws['A4'] = 'Total Entries:'
        ws['B4'] = sum(len(t.get('entries', [])) for t in tracker_data.values())
        
        ws['A5'] = 'Average Completion:'
        ws['B5'] = 75.5
        
        ws['A6'] = 'Current Month:'
        ws['B6'] = datetime.now().strftime('%B %Y')
        
        for col in range(1, 3):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_all_habits_data_sheet(self, ws, tracker_data):
        """Create all habits data sheet"""
        ws['A1'] = 'All Habits'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        headers = ['Habit', 'Entries', 'Streak', 'Completion %', 'Last Entry']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_statistics_data_sheet(self, ws, tracker_data):
        """Create statistics data sheet"""
        ws['A1'] = 'Detailed Statistics'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        headers = ['Habit', 'Total Entries', 'Average', 'Max', 'Min', 'Completion %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_insights_sheet(self, ws):
        """Create insights sheet"""
        ws['A1'] = 'Insights & Recommendations'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        insights = [
            'Keep up the great work!',
            'Try to maintain consistency in daily habits',
            'Focus on improving morning routine',
            'You are doing excellent with evening habits',
            'Consider setting reminders for afternoon check-ins'
        ]
        
        for idx, insight in enumerate(insights, 3):
            ws[f'A{idx}'] = f'• {insight}'
        
        ws.column_dimensions['A'].width = 50
    
    def _create_weekly_summary(self, ws):
        """Create weekly summary"""
        ws['A1'] = 'Weekly Summary'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['Day', 'Habits Completed', 'Total Target', 'Completion %']
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for row, day in enumerate(days, 4):
            ws.cell(row=row, column=1).value = day
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_weekly_details(self, ws):
        """Create weekly details"""
        ws['A1'] = 'Daily Details'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        headers = ['Date', 'Habit', 'Status', 'Value', 'Time']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_performance_sheet(self, ws):
        """Create performance sheet"""
        ws['A1'] = 'Performance Matrix'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        headers = ['Habit', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    def _create_goals_progress_sheet(self, ws):
        """Create goals progress sheet"""
        ws['A1'] = 'Goals Progress'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['Goal', 'Target', 'Achieved', 'Progress %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_monthly_summary(self, ws):
        """Create monthly summary"""
        ws['A1'] = 'Monthly Summary'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['Week', 'Completion %', 'Streak', 'Total Entries']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
        for row, week in enumerate(weeks, 4):
            ws.cell(row=row, column=1).value = week
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_monthly_daily_log(self, ws):
        """Create monthly daily log"""
        ws['A1'] = 'Daily Log - Full Month'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['Date', 'Habit', 'Value', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_monthly_statistics(self, ws):
        """Create monthly statistics"""
        ws['A1'] = 'Monthly Statistics'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        headers = ['Habit', 'Days Done', 'Success Rate', 'Best Day', 'Lowest Day']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_trends_sheet(self, ws):
        """Create trends sheet"""
        ws['A1'] = 'Trends Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        headers = ['Week', 'Trend', 'Change %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_achievements_sheet(self, ws):
        """Create achievements sheet"""
        ws['A1'] = 'Achievements'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        achievements = [
            '✓ Completed all habits for 1 week',
            '✓ Maintained 30-day streak',
            '✓ Reached 100 total entries',
            '✓ Perfect month (30/30 days)',
            '✓ Consistency master - no missed days'
        ]
        
        for idx, achievement in enumerate(achievements, 3):
            cell = ws.cell(row=idx, column=1)
            cell.value = achievement
            cell.font = Font(size=11)
        
        ws.column_dimensions['A'].width = 40
    
    def list_generated_files(self):
        """List all generated Excel files"""
        if not os.path.exists(self.output_dir):
            return []
        
        files = []
        for filename in os.listdir(self.output_dir):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(self.output_dir, filename)
                filesize = os.path.getsize(filepath)
                files.append({
                    'name': filename,
                    'path': filepath,
                    'size': filesize,
                    'url': f'{self.output_dir}/{filename}'
                })
        return files
